import json
import sys
from pathlib import Path

from openpyxl import load_workbook

BASE = Path(__file__).resolve().parent.parent
OUT = BASE / "scripts" / "xlsx-dump"
OUT.mkdir(parents=True, exist_ok=True)

FILES = [
    "洪涝阶段一.xlsx",
    "洪涝阶段二.xlsx",
    "保险方案综合评分计算表.xlsx",
    "灾情数据成本动因转换计算表.xlsx",
]


def cell_text(value):
    if value is None:
        return ""
    return str(value)


def dump(path: Path):
    result = {"file": path.name, "sheets": []}
    for data_only in (False, True):
        wb = load_workbook(path, data_only=data_only)
        for ws in wb.worksheets:
            rows = []
            for row in ws.iter_rows():
                rows.append([cell_text(c.value) for c in row])
            # trim trailing empty rows/cols
            while rows and not any(x.strip() for x in rows[-1]):
                rows.pop()
            entry = {
                "name": ws.title,
                "dims": ws.dimensions,
                "merged": [str(r) for r in ws.merged_cells.ranges],
                "rows": rows,
            }
            if not data_only:
                result["sheets"].append({"formulas": entry})
            else:
                idx = [i for i, s in enumerate(result["sheets"])]
            if data_only:
                pass
        if data_only:
            # attach values pass
            wb2 = load_workbook(path, data_only=True)
            for i, ws in enumerate(wb2.worksheets):
                rows = []
                for row in ws.iter_rows():
                    rows.append([cell_text(c.value) for c in row])
                while rows and not any(x.strip() for x in rows[-1]):
                    rows.pop()
                result["sheets"][i]["values"] = {"name": ws.title, "rows": rows}
    return result


def to_markdown(result):
    lines = [f"# {result['file']}", ""]
    for sheet in result["sheets"]:
        f = sheet["formulas"]
        v = sheet.get("values", {})
        lines.append(f"## Sheet: {f['name']}  (dims {f['dims']})")
        if f["merged"]:
            lines.append(f"merged: {', '.join(f['merged'])}")
        lines.append("")
        vrows = v.get("rows", [])
        for r_i, row in enumerate(f["rows"], start=1):
            vrow = vrows[r_i - 1] if r_i - 1 < len(vrows) else []
            cells = []
            for c_i, cell in enumerate(row):
                val = vrow[c_i] if c_i < len(vrow) else ""
                if cell.startswith("="):
                    cells.append(f"[{cell} => {val}]")
                elif cell.strip():
                    cells.append(cell)
                else:
                    cells.append("")
            while cells and not cells[-1].strip():
                cells.pop()
            if not cells:
                lines.append(f"R{r_i}:")
                continue
            lines.append(f"R{r_i}: " + " | ".join(cells))
        lines.append("")
    return "\n".join(lines)


for name in FILES:
    p = BASE / name
    if not p.exists():
        print(f"MISSING {p}", file=sys.stderr)
        continue
    res = dump(p)
    (OUT / (p.stem + ".json")).write_text(
        json.dumps(res, ensure_ascii=False, indent=1), encoding="utf-8"
    )
    (OUT / (p.stem + ".md")).write_text(to_markdown(res), encoding="utf-8")
    total = sum(len(s["formulas"]["rows"]) for s in res["sheets"])
    print(f"{p.name}: {len(res['sheets'])} sheets, {total} rows")
