"""Parse the two stage workbooks into the structured task content consumed by the app.

Run: python scripts/build_tasks.py
Output: src/data/generated/taskContent.js
"""

import json
import re
from pathlib import Path

from openpyxl import load_workbook

BASE = Path(__file__).resolve().parent.parent
OUT = BASE / "src" / "data" / "generated" / "taskContent.js"

STEP_RE = re.compile(r"^第([一二三四五六七八九十百]+)步[：:]\s*(.*)$")
ROLE_MARK_RE = re.compile(r"^【(.+?)】[：:]?\s*$")
BRACKET_ONLY_RE = re.compile(r"^【(.+?)】[。]?$")
CLICK_LEAD_RE = re.compile(r"^点击[：:]?$")
CLICK_INLINE_RE = re.compile(r"^点击\s*【(.+?)】[。]?$")
FIELD_RE = re.compile(r"^([^：:。，、（）]{2,14})[：:](.{1,60})$")
LABEL_RE = re.compile(r"^([^。，、]{2,20})[：:]$")
SUB_RE = re.compile(r"^((?:数据源|规则|合规项|方案|指标|字段|维度)\d*[^。]{0,24})[：:]?$")
OUTPUT_RE = re.compile(r"^输出[：:](.+)$")
DOC_RE = re.compile(r"《(.+?)》")

# Authoring notes left in the workbook for whoever builds the platform.
NOTE_MARKERS = ("弄到平台", "点完平台", "学生可以下载", "平台呈现效果")

ROLE_BY_SHEET = {
    "财务主管统筹岗": "finance-lead",
    "采购成本保障岗": "procurement",
    "应急预算绩效岗": "budget-performance",
    "资金核算风控岗": "fund-risk",
}

CN_NUM = "一二三四五六七八九十"


def read_sheet(path, sheet):
    wb = load_workbook(BASE / path, data_only=True)
    ws = wb[sheet]
    rows = {}
    for row in ws.iter_rows():
        cells = [("" if c.value is None else str(c.value).strip()) for c in row]
        while cells and not cells[-1]:
            cells.pop()
        while cells and not cells[0]:
            cells.pop(0)
        if cells:
            rows[row[0].row] = cells
    return rows


def slice_rows(rows, start, end):
    return [(i, rows[i]) for i in range(start, end + 1) if i in rows]


def is_path(text):
    return "→" in text and "。" not in text and len(text) < 80


def is_formula(text):
    if "=" not in text:
        return False
    if text.startswith("公式") or text.startswith("统一公式"):
        return True
    head = text.split("=")[0]
    return len(head) < 40 and ("。" not in head)


def is_quote(text):
    return text.startswith("“") or text.startswith('"')


def is_note(text):
    return any(marker in text for marker in NOTE_MARKERS)


def classify_single(text):
    text = text.strip()
    if is_note(text):
        return {"t": "note", "v": text}
    return {"t": "text", "v": text}


class Collector:
    """Turns a run of raw worksheet lines into typed render blocks."""

    def __init__(self):
        self.blocks = []
        self._fields = []
        self._list = []
        self._code = []
        self._table = []

    def _flush_fields(self):
        if self._fields:
            self.blocks.append({"t": "fields", "items": self._fields})
            self._fields = []

    def _flush_list(self):
        if self._list:
            if len(self._list) == 1:
                self.blocks.append({"t": "text", "v": self._list[0]})
            else:
                self.blocks.append({"t": "list", "items": self._list})
            self._list = []

    def _flush_code(self):
        if self._code:
            while self._code and not self._code[-1].strip():
                self._code.pop()
            self.blocks.append({"t": "code", "lines": self._code})
            self._code = []

    def _flush_table(self):
        if not self._table:
            return
        rows = self._table
        self._table = []
        width = max(len(r) for r in rows)
        rows = [r + [""] * (width - len(r)) for r in rows]
        keep = [c for c in range(width) if any(r[c].strip() for r in rows)]
        rows = [[r[c] for c in keep] for r in rows]
        # A single-cell "table" is really just a stray note or caption.
        if len(rows) == 1 and len(rows[0]) == 1:
            self.blocks.append(classify_single(rows[0][0]))
            return
        head, *body = rows
        self.blocks.append({"t": "table", "head": head, "rows": body})

    def flush(self):
        self._flush_code()
        self._flush_table()
        self._flush_fields()
        self._flush_list()

    def add(self, block):
        self.flush()
        self.blocks.append(block)

    def field(self, label, value):
        self._flush_code()
        self._flush_table()
        self._flush_list()
        self._fields.append([label, value])

    def bullet(self, text):
        self._flush_code()
        self._flush_table()
        self._flush_fields()
        self._list.append(text)

    def code(self, text):
        self._flush_table()
        self._flush_fields()
        self._flush_list()
        self._code.append(text)

    def table(self, cells):
        self._flush_code()
        self._flush_fields()
        self._flush_list()
        self._table.append(cells)

    def result(self):
        self.flush()
        return self.blocks


def parse_blocks(lines, outputs):
    """lines: list of cell-lists (already row-sliced)."""
    col = Collector()
    pending_speech = False
    in_code = False
    i = 0
    while i < len(lines):
        cells = lines[i]
        text = cells[0].strip()
        i += 1

        if len(cells) > 1:
            in_code = False
            pending_speech = False
            col.table(cells)
            continue

        if not text:
            continue

        # Python source block inside stage-1 task 2.
        if text.startswith("import ") or text.startswith("from "):
            in_code = True
        if in_code:
            if text.startswith("点击") or text.startswith("系统") or text.startswith("平台"):
                in_code = False
            else:
                col.code(text)
                continue

        out = OUTPUT_RE.match(text)
        if out:
            docs = DOC_RE.findall(out.group(1))
            outputs.extend(docs if docs else [out.group(1).strip("。")])
            continue

        if text in ("现场口播", "现场口播：", "此时进行现场口播：", "口播", "口播："):
            pending_speech = True
            continue

        if is_note(text):
            col.add({"t": "note", "v": text})
            continue

        marker = ROLE_MARK_RE.match(text)
        if marker and ("口播" in marker.group(1) or "岗" in marker.group(1)):
            if "口播" in marker.group(1):
                pending_speech = True
            continue

        if is_quote(text):
            col.add({"t": "speech", "v": text.strip("“”\"")})
            pending_speech = False
            continue

        if pending_speech:
            col.add({"t": "speech", "v": text.strip("“”\"")})
            pending_speech = False
            continue

        inline_click = CLICK_INLINE_RE.match(text)
        if inline_click:
            col.add({"t": "action", "v": inline_click.group(1)})
            continue

        if CLICK_LEAD_RE.match(text):
            if i < len(lines):
                nxt = lines[i][0].strip()
                bracket = BRACKET_ONLY_RE.match(nxt)
                if bracket:
                    col.add({"t": "action", "v": bracket.group(1)})
                    i += 1
                    continue
            continue

        bracket = BRACKET_ONLY_RE.match(text)
        if bracket:
            col.add({"t": "action", "v": bracket.group(1)})
            continue

        if is_path(text):
            col.add({"t": "path", "v": text})
            continue

        if is_formula(text):
            col.add({"t": "formula", "v": text})
            continue

        field = FIELD_RE.match(text)
        if field and "。" not in field.group(2) and "→" not in text:
            col.field(field.group(1), field.group(2).strip())
            continue

        label = LABEL_RE.match(text)
        if label:
            # "进入：" / "设置：" only introduce the next line; the block that
            # follows already carries the meaning, so drop the empty lead-in.
            if label.group(1) in ("进入", "设置", "选择", "建立", "新增", "包括", "返回"):
                continue
            sub = SUB_RE.match(text)
            if sub:
                col.add({"t": "sub", "v": label.group(1)})
            else:
                col.add({"t": "label", "v": label.group(1)})
            continue

        if len(text) <= 26 and "。" not in text and "，" not in text:
            col.bullet(text)
            continue

        col.add({"t": "text", "v": text})

    return col.result()


def split_step_title(rest):
    """Stage 2 packs the step title and its first sentence onto one line."""
    if "。" in rest:
        head, _, tail = rest.partition("。")
        if len(head) <= 28:
            return head, ([[tail.strip()]] if tail.strip() else [])
        rest = head + "。" + tail
    if len(rest) > 28 and "，" in rest:
        head, _, tail = rest.partition("，")
        return head, ([[tail.strip().rstrip("。") + "。"]] if tail.strip() else [])
    return rest.rstrip("。"), []


def split_steps(entries):
    """Split a row run on 第N步 markers. Returns [(label, title, lines)]."""
    groups = []
    current = ["", "", []]
    for _, cells in entries:
        head = cells[0].strip()
        m = STEP_RE.match(head) if len(cells) == 1 else None
        if m:
            if current[2] or current[0]:
                groups.append(current)
            title, body = split_step_title(m.group(2).strip())
            current = [f"第{m.group(1)}步", title, body]
        else:
            current[2].append(cells)
    if current[2] or current[0]:
        groups.append(current)
    return groups


def build_step(step_id, role_id, label, title, lines, outputs, sub_role=None):
    blocks = parse_blocks(lines, outputs)
    step = {"id": step_id, "roleId": role_id, "label": label, "title": title, "blocks": blocks}
    if sub_role:
        step["subRole"] = sub_role
    return step


# --------------------------------------------------------------------------
# Task configuration. Row ranges are 1-based and inclusive, matching the
# worksheets exactly; see scripts/xlsx-compact/*.txt for the raw dumps.
# --------------------------------------------------------------------------

STAGE1_FILE = "洪涝阶段一.xlsx"
STAGE1_FULL_FILE = "洪涝阶段一全.xlsx"
STAGE2_FILE = "洪涝阶段二.xlsx"

# 补充表工作表名带有“从第N行开始”，读入后按岗位短名归档，便于与阶段二同一套分段逻辑复用。
STAGE1_FULL_SHEETS = {
    "财务主管统筹岗": "财务主管统筹岗-1020行开始",
    "采购成本保障岗": "采购成本保障岗-656行开始",
    "应急预算绩效岗": "应急预算绩效岗-57行开始",
    "资金核算风控岗": "资金核算风控岗-1行",
}

STAGE1 = [
    {
        "key": "s1-t1",
        "no": 1,
        "title": "启用洪涝应急救援专项账套",
        "owner": "finance-lead",
        "sheet": "财务主管统筹岗",
        "rows": (2, 367),
        "summary": "建立洪涝应急救援专项账套，配置专项科目、资金来源、9网格辅助核算、岗位权限与5项内控规则，打通“需求—物资—资金—凭证”四链联动。",
        "panel": "ledger",
        "outputs": ["专项账套配置校验单", "岗位权限配置表", "内控规则清单"],
    },
    {
        "key": "s1-t2",
        "no": 2,
        "title": "建设洪涝应急救援数据采集系统",
        "owner": "procurement",
        "sheet": "采购成本保障岗",
        "rows": (2, 604),
        "summary": "接入应急管理、气象、无人机、御洪星四类数据源，建立统一字段模型与 Python 采集节点，完成 9 网格数据采集、敏感数据分级与 IQR 异常监测。",
        "panel": "collect",
        "outputs": ["洪涝应急救援九网格原始数据表", "数据字段映射表", "IQR异常监测规则"],
    },
    {
        "key": "s1-t3",
        "no": 3,
        "title": "数据建模分析合规性检测",
        "owner": "finance-lead",
        "sheet": "财务主管统筹岗",
        "rows": (369, 1017),
        "summary": "对清洗后的 9 网格数据执行两层校验：第一层 95% 内部质量门槛，第二层 6 项合规一票否决，判定是否准予进入预算模型。",
        "panel": "compliance",
        "outputs": ["灾情数据质量与合规校验单", "模型准入判定结果"],
    },
    {
        "key": "s1-t4",
        "no": 4,
        "title": "救援人员保险方案比较",
        "owner": "procurement",
        "sheet": "采购成本保障岗",
        "rows": (606, 653),
        "summary": "对 A/B/C 三家保险产品按 8 项指标加权评分，成本型与效益型指标分别标准化，选出综合得分最高方案并测算总保费。",
        "panel": "insurance",
        "steps": [
            ("导入三家保险产品报价", 608, 621),
            ("设置指标权重与承保范围分档", 624, 632),
            ("确定标准化规则", 634, 636),
            ("下载评分计算表并导入平台", 638, 645),
            ("输出方案选择结论", 647, 653),
        ],
        "outputs": ["保险方案综合评分计算表", "保险方案选择结论"],
    },
    {
        "key": "s1-t5",
        "no": 5,
        "title": "将灾情数据转换为成本动因",
        "owner": "budget-performance",
        "sheet": "应急预算绩效岗",
        "rows": (1, 48),
        "summary": "把 9 网格灾情数据按 7 类核心公式转换为预算计算参数，形成网格预算、保险预算与设备预算，汇总总预算需求。",
        "panel": "cost-driver",
        "steps": [
            ("明确成本动因转换公式", 2, 11),
            ("载入历史采购价参数", 13, 26),
            ("下载成本动因计算表并导入平台", 29, 35),
            ("输出总预算需求", 37, 37),
        ],
        "outputs": ["灾情数据成本动因转换计算表", "9网格预算测算表"],
    },
]

# 洪涝阶段一全.xlsx 在原 5 个任务之后续写的内容。任务 6、8 在表中有标题；
# 「B方案预算审批」夹在任务 6 与任务 8 之间，按执行顺序编为任务 7。
STAGE1_EXTRA = [
    {
        "key": "s1-t6",
        "no": 6,
        "title": "编制ABC等级预算",
        "owner": "budget-performance",
        "summary": "按轻度、中度、重度编制 A/B/C 三套预算：A 方案 2,816,906 元、B 方案 2,909,004 元、C 方案 4,278,517.50 元，并形成单位受益成本与适用灾情对照。",
        "panel": "abc-budget",
        "segments": [("应急预算绩效岗", 59, 82, "编制三受灾等级预算")],
        "outputs": ["ABC三受灾等级预算计算表", "ABC三方案预算"],
    },
    {
        "key": "s1-t7",
        "no": 7,
        "title": "B方案预算审批",
        "owner": "finance-lead",
        "summary": "登记 III 级响应并载入 B 方案，按已确认到账 366 万元测算资金覆盖率 125.82%，建立 2,909,004 元预算控制额度并完成审批同步。",
        "panel": "budget-approval",
        "segments": [("财务主管统筹岗", 1022, 1257, None)],
        "outputs": ["B方案应急预算审批单", "资金保障测算表", "预备费控制台账"],
    },
    {
        "key": "s1-t8",
        "no": 8,
        "title": "第一次突发事件——受灾人数突然增加",
        "owner": "budget-performance",
        "summary": "甲3、甲6二次报送后响应由 III 级升为 II 级，预算由 B 方案切换为 C 方案；财政资金全部到位后覆盖率 93.96%，短期缺口 258,517.50 元，并启动帐篷追加采购。",
        "panel": "emergency-update",
        "segments": [
            ("应急预算绩效岗", 86, 166, None),
            ("财务主管统筹岗", 1261, 1270, "预算方案二次决策"),
            ("采购成本保障岗", 659, 679, "新增物资分析"),
            ("资金核算风控岗", 3, 17, "资金状态汇总"),
        ],
        "outputs": ["二次灾情数据更新表", "C方案预算参数", "资金韧性测算"],
    },
]

STAGE2 = [
    {
        "key": "s2-t1",
        "no": 1,
        "title": "生成9网格采购需求",
        "owner": "procurement",
        "summary": "按安置方式与帐篷可用量测算 6 类物资净采购量，划分合同采购与生活保障直采，并由财务主管复核后生成采购任务。",
        "panel": "demand",
        "segments": [
            ("采购成本保障岗", 2, 94, None),
            ("财务主管统筹岗", 2, 26, "采购需求审核"),
        ],
        "outputs": ["9网格采购需求测算表", "采购需求复核单"],
    },
    {
        "key": "s2-t2",
        "no": 2,
        "title": "建立分层采购价格基准",
        "owner": "procurement",
        "summary": "接收供应商报价后下载分层采购价格计算表，导入测算结果，并按 5%/10% 设置绿黄红偏差预警。",
        "panel": "price",
        "segments": [("采购成本保障岗", 96, 145, None)],
        "outputs": [
            "分层采购价格基准计算表",
            "4类合同物资价格基准表",
            "2类生活保障物资应急零售/框架直采价格核验表",
            "价格偏差分析表",
            "报价口径校验单",
        ],
    },
    {
        "key": "s2-t3",
        "no": 3,
        "title": "合同物资供应商综合评分与初始遴选",
        "owner": "procurement",
        "summary": "按报价 40%、交付 20%、质量 15%、资质 10%、履约 10%、距离 5% 加权评分，形成 S2 > S1 > S3 初始排序并完成三岗复核。",
        "panel": "supplier",
        "segments": [
            ("采购成本保障岗", 147, 176, "综合评分与初始排序"),
            ("应急预算绩效岗", 2, 15, "预算和网格保障复核"),
            ("资金核算风控岗", 2, 12, "核验S2相关资质"),
            ("财务主管统筹岗", 31, 48, "审批S2作为供应商"),
        ],
        "outputs": ["供应商综合评分表", "初始遴选结论"],
    },
    {
        "key": "s2-t4",
        "no": 4,
        "title": "初始合同、直采控制与预算占用",
        "owner": "procurement",
        "summary": "编制 HT-2025-001 合同 932,460 元与生活保障直采协议 149,397.50 元，占用 C 方案预算 25.29%，并完成资金控制与履约监测分流。",
        "panel": "contract",
        "segments": [
            ("采购成本保障岗", 178, 317, "合同与直采协议"),
            ("应急预算绩效岗", 20, 31, "C方案预算占用率测算"),
            ("资金核算风控岗", 17, 86, "采购金额核对与资金控制"),
            ("财务主管统筹岗", 53, 211, "审批合同采购和生活保障直采方案"),
        ],
        "outputs": ["HT-2025-001采购合同", "生活保障直采协议", "C方案预算占用率测算表", "采购资金台账"],
    },
    {
        "key": "s2-t5",
        "no": 5,
        "title": "第二次突发事件——供应商库存突变，重点物资无法按时足量交付",
        "owner": "procurement",
        "summary": "交叉核验 S2 仓库局部进水属实后，测算 12 小时缺口 200 顶、最终合同供应缺口 150 顶，形成可复制到 Excel 的影响测算表。",
        "panel": "split",
        "segments": [("采购成本保障岗", 324, 362, None)],
        "outputs": ["异常真实性核验结果", "合同影响测算表"],
    },
]


def build_stage1(sheets):
    tasks = []
    for cfg in STAGE1:
        rows = sheets[cfg["sheet"]]
        outputs = list(cfg.get("outputs", []))
        steps = []
        if "steps" in cfg:
            for idx, (title, a, b) in enumerate(cfg["steps"], start=1):
                lines = [cells for _, cells in slice_rows(rows, a, b)]
                steps.append(
                    build_step(
                        f"{cfg['key']}-s{idx}",
                        cfg["owner"],
                        f"第{CN_NUM[idx - 1]}步",
                        title,
                        lines,
                        outputs,
                    )
                )
        else:
            entries = slice_rows(rows, *cfg["rows"])
            # Drop the "任务N：..." heading row.
            entries = [e for e in entries if not e[1][0].startswith("任务")]
            groups = split_steps(entries)
            idx = 0
            for label, title, lines in groups:
                if not label:
                    blocks = parse_blocks(lines, outputs)
                    if blocks:
                        steps.append(
                            {
                                "id": f"{cfg['key']}-s0",
                                "roleId": cfg["owner"],
                                "label": "任务导入",
                                "title": "任务背景与目标",
                                "blocks": blocks,
                            }
                        )
                    continue
                idx += 1
                steps.append(
                    build_step(f"{cfg['key']}-s{idx}", cfg["owner"], label, title, lines, outputs)
                )

        tasks.append(
            {
                "key": cfg["key"],
                "stage": 1,
                "no": cfg["no"],
                "title": cfg["title"],
                "owner": cfg["owner"],
                "roles": [cfg["owner"]],
                "summary": cfg["summary"],
                "panel": cfg.get("panel"),
                "source": f"洪涝阶段一.xlsx · {cfg['sheet']}",
                "outputs": dedupe(outputs),
                "steps": steps,
            }
        )
    return tasks


def build_stage2(sheets, configs=None, stage=2, source="洪涝阶段二.xlsx"):
    tasks = []
    for cfg in configs or STAGE2:
        outputs = list(cfg.get("outputs", []))
        steps = []
        roles = []

        if "outputRows" in cfg:
            sheet, a, b = cfg["outputRows"]
            parse_blocks([cells for _, cells in slice_rows(sheets[sheet], a, b)], outputs)

        intro_blocks = []
        if "intro" in cfg:
            sheet, a, b = cfg["intro"]
            lines = [cells for _, cells in slice_rows(sheets[sheet], a, b)]
            lines = [c for c in lines if not c[0].startswith("任务")]
            intro_blocks = parse_blocks(lines, outputs)
        if intro_blocks:
            steps.append(
                {
                    "id": f"{cfg['key']}-s0",
                    "roleId": cfg["owner"],
                    "label": "任务导入",
                    "title": "任务背景与目标",
                    "blocks": intro_blocks,
                }
            )

        if "segments" in cfg:
            for sheet, a, b, seg_title in cfg["segments"]:
                role_id = ROLE_BY_SHEET[sheet]
                if role_id not in roles:
                    roles.append(role_id)
                entries = slice_rows(sheets[sheet], a, b)
                entries = [e for e in entries if not e[1][0].startswith("任务")]
                groups = split_steps(entries)
                has_steps = any(g[0] for g in groups)
                if has_steps:
                    idx = 0
                    for label, title, lines in groups:
                        if not label:
                            blocks = parse_blocks(lines, outputs)
                            if blocks:
                                steps.append(
                                    {
                                        "id": f"{cfg['key']}-{role_id}-intro",
                                        "roleId": role_id,
                                        "label": "任务导入",
                                        "title": seg_title or "任务背景与目标",
                                        "blocks": blocks,
                                    }
                                )
                            continue
                        idx += 1
                        steps.append(
                            build_step(
                                f"{cfg['key']}-{role_id}-{idx}", role_id, label, title, lines, outputs
                            )
                        )
                else:
                    lines = [cells for _, cells in entries]
                    steps.append(
                        build_step(
                            f"{cfg['key']}-{role_id}",
                            role_id,
                            sheet,
                            seg_title or "岗位处置",
                            lines,
                            outputs,
                        )
                    )

        if "steps2" in cfg:
            for idx, item in enumerate(cfg["steps2"], start=1):
                role_id, sheet, a, b, override = item[:5]
                sub_role = item[5] if len(item) > 5 else None
                if role_id not in roles:
                    roles.append(role_id)
                entries = slice_rows(sheets[sheet], a, b)
                groups = split_steps(entries)
                label, title, lines = groups[0]
                if override:
                    label, title = override
                if not label:
                    label = f"第{idx}步"
                steps.append(
                    build_step(
                        f"{cfg['key']}-s{idx}", role_id, label, title, lines, outputs, sub_role
                    )
                )

        if not roles:
            roles = [cfg["owner"]]

        tasks.append(
            {
                "key": cfg["key"],
                "stage": stage,
                "no": cfg["no"],
                "title": cfg["title"],
                "owner": cfg["owner"],
                "roles": roles,
                "summary": cfg["summary"],
                "panel": cfg.get("panel"),
                "source": source,
                "outputs": dedupe(outputs),
                "steps": steps,
            }
        )
    return tasks


def dedupe(items):
    seen, result = set(), []
    for item in items:
        if item not in seen:
            seen.add(item)
            result.append(item)
    return result


def main():
    role_sheets = tuple(ROLE_BY_SHEET)
    s1_sheets = {name: read_sheet(STAGE1_FILE, name) for name in role_sheets}
    s1_full = {name: read_sheet(STAGE1_FULL_FILE, sheet) for name, sheet in STAGE1_FULL_SHEETS.items()}
    s2_sheets = {name: read_sheet(STAGE2_FILE, name) for name in role_sheets}

    extra = build_stage2(s1_full, STAGE1_EXTRA, stage=1, source="洪涝阶段一全.xlsx")
    tasks = build_stage1(s1_sheets) + extra + build_stage2(s2_sheets)

    for task in tasks:
        for step in task["steps"]:
            if step["roleId"] not in task["roles"]:
                task["roles"].append(step["roleId"])

    payload = json.dumps(tasks, ensure_ascii=False, indent=2)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(
        "// 由 scripts/build_tasks.py 从洪涝阶段一/阶段一全/阶段二工作簿生成，请勿手工编辑。\n"
        "// 重新生成：python scripts/build_tasks.py\n\n"
        f"export const taskContent = {payload}\n",
        encoding="utf-8",
    )

    print(f"tasks: {len(tasks)}")
    for task in tasks:
        blocks = sum(len(s["blocks"]) for s in task["steps"])
        print(
            f"  {task['key']} S{task['stage']}-{task['no']} {task['title']} "
            f"steps={len(task['steps'])} blocks={blocks} roles={','.join(task['roles'])} "
            f"outputs={len(task['outputs'])}"
        )


if __name__ == "__main__":
    main()
